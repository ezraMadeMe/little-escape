#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
서울맛집.xlsx 데이터를 읽어서 Place 테이블에 삽입할 SQL 생성
"""

import openpyxl
from pathlib import Path

# Excel 파일 경로
excel_path = Path(__file__).parent.parent / "data" / "서울맛집.xlsx"
output_sql_path = Path(__file__).parent.parent / "data" / "insert_seoul_restaurants.sql"

print(f"Excel 파일 읽기: {excel_path}")

# Excel 파일 열기
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# 헤더는 5번째 행
header_row = 5
headers = []
for col in range(1, sheet.max_column + 1):
    header = sheet.cell(header_row, col).value
    headers.append(header)

print(f"헤더: {headers}")
print(f"총 데이터 행: {sheet.max_row - header_row}")

# 데이터 파싱
restaurants = []
for row in range(header_row + 1, sheet.max_row + 1):
    row_data = []
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row, col).value
        row_data.append(value if value else "")

    # 빈 행 스킵
    if not any(row_data):
        continue

    # 리스트로 저장
    restaurants.append(row_data)

workbook.close()

print(f"\n총 {len(restaurants)}개의 맛집 데이터 파싱 완료")

# 첫 3개 샘플 출력
print("\n=== 데이터 샘플 (첫 3개) ===")
for i, restaurant in enumerate(restaurants[:3], start=1):
    print(f"\n{i}. {restaurant}")

# SQL INSERT 문 생성
sql_statements = []
sql_statements.append("-- 서울 맛집 데이터 삽입")
sql_statements.append("-- Generated from 서울맛집.xlsx")
sql_statements.append("")

valid_count = 0
for i, row_data in enumerate(restaurants, start=1):
    # 인덱스: 0=시도, 1=시군구, 2=음식점명, 3=대표메뉴, 4=전화번호, 5=네이버 검색명, 6=추천사유
    if len(row_data) < 7:
        print(f"경고: {i}번 행 - 데이터가 부족하여 스킵 (길이: {len(row_data)})")
        continue

    sido = str(row_data[0]).strip() if row_data[0] else ""
    sigungu = str(row_data[1]).strip() if row_data[1] else ""
    restaurant_name = str(row_data[2]).strip() if row_data[2] else ""
    main_menu = str(row_data[3]).strip() if row_data[3] else ""
    phone = str(row_data[4]).strip() if row_data[4] else ""
    search_keyword = str(row_data[5]).strip() if row_data[5] else ""
    recommendation = str(row_data[6]).strip() if row_data[6] else ""

    # 유효성 검사 - 음식점명 또는 검색 키워드가 있어야 함
    if not restaurant_name and not search_keyword:
        print(f"경고: {i}번 행 - 음식점명과 검색명이 모두 없어 스킵")
        continue

    # 주소 생성
    address = f"{sido} {sigungu}".strip()
    if not address:
        address = "서울특별시"

    # 이름: 검색 키워드 > 음식점명
    name = search_keyword if search_keyword else restaurant_name

    # URL은 네이버 검색 링크로 생성
    search_term = search_keyword if search_keyword else restaurant_name
    url = f"https://search.naver.com/search.naver?query={search_term}"

    # 위도/경도는 임시로 서울 중심 좌표 사용
    latitude = 37.5665
    longitude = 126.9780

    # SQL Injection 방지를 위해 작은따옴표 이스케이프
    name_safe = name.replace("'", "''")
    address_safe = address.replace("'", "''")

    sql = f"""INSERT INTO places (name, address, url, latitude, longitude, category, image_url, created_at, updated_at)
VALUES ('{name_safe}', '{address_safe}', '{url}', {latitude}, {longitude}, 'FOOD', NULL, NOW(), NOW());"""

    sql_statements.append(sql)
    valid_count += 1

# SQL 파일 저장
with open(output_sql_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_statements))

print(f"\nSQL 파일 생성 완료: {output_sql_path}")
print(f"총 {valid_count}개의 INSERT 문 생성")
print("\n주의: 위도/경도는 서울 중심 좌표(37.5665, 126.9780)로 임시 설정되었습니다.")
print("실제 사용하려면 주소 -> 좌표 변환(지오코딩)이 필요합니다.")
